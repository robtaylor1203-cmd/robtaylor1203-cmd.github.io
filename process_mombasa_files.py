import os
import json
import datetime
import zipfile
import shutil
import fitz  # PyMuPDF for PDFs
import docx  # python-docx for .docx files
import openpyxl  # openpyxl for .xlsx files

# --- Configuration ---
BASE_PATH = os.getcwd() 
INBOX_PATH = os.path.join(BASE_PATH, "Inbox", "Momabsa")
OUTPUT_DIR = os.path.join(BASE_PATH, "source_reports", "mombasa_sale_reports")

# --- Reference data for calculating the sale number ---
REFERENCE_DATE = datetime.date(2025, 9, 1)
REFERENCE_SALE_NO = 34

def calculate_upcoming_sale_no():
    """
    Calculates the sale number for the upcoming Monday.
    """
    today = datetime.date.today()
    days_until_next_monday = (7 - today.weekday()) % 7
    if days_until_next_monday == 0:
        days_until_next_monday = 7
        
    upcoming_monday = today + datetime.timedelta(days=days_until_next_monday)
    weeks_passed = (upcoming_monday - REFERENCE_DATE).days // 7
    upcoming_sale_no = REFERENCE_SALE_NO + weeks_passed
    
    print(f"Today's Date: {today}")
    print(f"Upcoming Monday: {upcoming_monday}")
    print(f"Calculated Upcoming Sale Number: {upcoming_sale_no}")
    
    return upcoming_sale_no

def extract_text_from_file(file_path):
    """
    Intelligently extracts text from a file based on its extension.
    """
    _, extension = os.path.splitext(file_path)
    extension = extension.lower()
    text = ""

    print(f"  - Reading file: {os.path.basename(file_path)}")

    if extension == ".txt":
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
    elif extension == ".pdf":
        with fitz.open(file_path) as doc:
            for page in doc:
                text += page.get_text()
    elif extension == ".docx":
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + '\n'
    elif extension == ".xlsx":
        workbook = openpyxl.load_workbook(file_path)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            text += f"--- Sheet: {sheetname} ---\n"
            for row in sheet.iter_rows(values_only=True):
                text += "\t".join(str(cell) for cell in row if cell is not None) + "\n"
    else:
        text = f"Unsupported file type: {extension}"
        print(f"    Warning: {text}")
        
    return text

def process_local_sale_files():
    """
    Main function to find and process local ZIP, TXT, and PDF files for a given sale.
    """
    print(f"--- Starting Local File Processor for Mombasa Sales ---")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    try:
        sale_no = calculate_upcoming_sale_no()
        
        # --- Define file and directory paths ---
        zip_filename = f"{sale_no}.zip"
        txt_filename = f"Mombasa Sale {sale_no}.txt"
        # --- NEW: Define the new PDF filename ---
        pdf_filename = f"Final Market Report Sale {sale_no} 2025.pdf"

        zip_filepath = os.path.join(INBOX_PATH, zip_filename)
        txt_filepath = os.path.join(INBOX_PATH, txt_filename)
        # --- NEW: Define the new PDF filepath ---
        pdf_filepath = os.path.join(INBOX_PATH, pdf_filename)
        
        temp_extract_path = os.path.join(INBOX_PATH, f"temp_{sale_no}")

        final_data = {
            "sale_number": sale_no,
            "retrieval_date": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "main_report_text": None,
            # --- NEW: Add a key for the new PDF's content ---
            "final_market_report_pdf_text": None,
            "zip_attachments_content": {}
        }

        # --- PART 1: Process the Main Text File ---
        print(f"\n[1/3] Processing main text file: {txt_filename}")
        if os.path.exists(txt_filepath):
            final_data["main_report_text"] = extract_text_from_file(txt_filepath)
            print("  -> Success.")
        else:
            print(f"  -> Warning: File not found at {txt_filepath}")
            final_data["main_report_text"] = "File not found."

        # --- NEW PART: Process the Final Market Report PDF ---
        print(f"\n[2/3] Processing Final Market Report PDF: {pdf_filename}")
        if os.path.exists(pdf_filepath):
            final_data["final_market_report_pdf_text"] = extract_text_from_file(pdf_filepath)
            print("  -> Success.")
        else:
            print(f"  -> Warning: File not found at {pdf_filepath}")
            final_data["final_market_report_pdf_text"] = "File not found."

        # --- PART 3: Process the ZIP File ---
        print(f"\n[3/3] Processing ZIP file: {zip_filename}")
        if os.path.exists(zip_filepath):
            os.makedirs(temp_extract_path, exist_ok=True)
            print(f"  - Extracting to temporary directory...")
            with zipfile.ZipFile(zip_filepath, 'r') as zip_ref:
                zip_ref.extractall(temp_extract_path)
            
            print(f"  - Reading contents of extracted files...")
            for filename in os.listdir(temp_extract_path):
                filepath = os.path.join(temp_extract_path, filename)
                if os.path.isfile(filepath):
                    content = extract_text_from_file(filepath)
                    final_data["zip_attachments_content"][filename] = content
            
            print("  - Cleaning up temporary files...")
            shutil.rmtree(temp_extract_path)
            print("  -> Success.")
        else:
            print(f"  -> Warning: ZIP file not found at {zip_filepath}")
            final_data["zip_attachments_content"] = {"error": "File not found."}

        # --- Save the combined data to a single JSON file ---
        output_filename = f"Mombasa_Sale_{sale_no}_Combined_{datetime.date.today()}.json"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        print(f"\nSaving all extracted data to: {output_path}")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(final_data, f, indent=4, ensure_ascii=False)

        print("\n--- PROCESS COMPLETE ---")

    except Exception as e:
        print(f"!!! An unexpected error occurred: {e}")

if __name__ == "__main__":
    process_local_sale_files()
